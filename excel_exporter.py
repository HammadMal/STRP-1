"""
Excel Export Module for CLO/PLO Mapping Tool
Handles creation of formatted Excel reports with student performance data.
Modified to append results to the original uploaded file instead of creating a new file.
FIXED: Only shows CLOs that actually exist in the data (no more hardcoded CLO 1-5)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import os


def create_excel_output(clo_scores, plo_scores, grades, data_dict, original_file_path):
    """
    Append CLO, PLO scores, and final grades to the original Excel file as a new sheet.
    Only includes CLOs that actually exist in the course structure.
    
    Args:
        clo_scores (dict): Dictionary of CLO scores for each student
        plo_scores (dict): Dictionary of PLO scores for each student
        grades (dict): Dictionary of final grades for each student (from terminal calculation)
        data_dict (dict): Original data dictionary from data.py
        original_file_path (str): Path to the original uploaded file
    
    Returns:
        str: Path to the updated Excel file (same as original)
    """
    
    # Check if the file is Excel format
    file_ext = os.path.splitext(original_file_path)[1].lower()
    if file_ext not in ['.xlsx', '.xls']:
        raise ValueError("Can only append to Excel files. Original file must be .xlsx or .xls format.")
    
    # Prepare data for DataFrame
    data_rows = []
    
    # Get CLOs that are actually defined in the course structure
    # Use the 'clos' dictionary which contains ALL defined CLOs (even without assessments)
    defined_clos = set()
    if 'clos' in data_dict and data_dict['clos']:
        defined_clos = set(data_dict['clos'].keys())
        print(f"🔍 CLOs found in course definition: {sorted(defined_clos)}")
    else:
        # Fallback: use clo_assessments (only CLOs with actual assessments)
        if 'clo_assessments' in data_dict and data_dict['clo_assessments']:
            defined_clos = set(data_dict['clo_assessments'].keys())
            print(f"⚠️ Using CLOs from assessments (fallback): {sorted(defined_clos)}")
        else:
            # Last resort: use what we have in scores
            for student_clos in clo_scores.values():
                defined_clos.update(student_clos.keys())
            print(f"⚠️ Using CLOs from scores (last resort): {sorted(defined_clos)}")
    
    # Get PLOs that are actually mapped in the course
    defined_plos = set()
    if 'clo_to_plo' in data_dict and data_dict['clo_to_plo']:
        for clo_mapping in data_dict['clo_to_plo'].values():
            if isinstance(clo_mapping, dict) and 'PLO' in clo_mapping:
                defined_plos.add(clo_mapping['PLO'])
        print(f"🔍 PLOs found in course mapping: {sorted(defined_plos)}")
    else:
        # Fallback: use what we have in scores
        for student_plos in plo_scores.values():
            defined_plos.update(student_plos.keys())
        print(f"⚠️ Using PLOs from scores (fallback): {sorted(defined_plos)}")
    
    # Remove CLO 0 if it exists (bonus points typically excluded from final display)
    if 'CLO 0' in defined_clos:
        defined_clos.discard('CLO 0')
        print(f"📝 Excluding CLO 0 from Excel output (bonus points)")
    
    # Sort CLOs and PLOs naturally
    sorted_clos = sorted(defined_clos, key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 999)
    sorted_plos = sorted(defined_plos, key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 999)
    
    print(f"📊 Final CLOs for Excel: {sorted_clos}")
    print(f"📊 Final PLOs for Excel: {sorted_plos}")
    
    # Create data rows
    for student_id in clo_scores.keys():
        row_data = {'ID': student_id}
        
        # Add CLO scores - ONLY the ones that are defined in the course
        for clo in sorted_clos:
            row_data[clo] = clo_scores[student_id].get(clo, 0)
        
        # Add PLO scores - ONLY the ones that are defined in the course
        for plo in sorted_plos:
            row_data[plo] = plo_scores[student_id].get(plo, 0)
        
        # Add overall grade (from terminal calculation)
        grade_percentage = grades.get(student_id, 0)
        letter_grade = _calculate_letter_grade(grade_percentage)
        row_data['Overall Grade'] = f"{grade_percentage:.2f}% ({letter_grade})"
        
        data_rows.append(row_data)
    
    # Create DataFrame
    df = pd.DataFrame(data_rows)
    
    # Load the existing workbook and append new sheet
    try:
        # Load existing workbook
        with pd.ExcelWriter(original_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Write main data to "CLO PLO Results" sheet
            df.to_excel(writer, sheet_name='CLO PLO Results', index=False)
            
            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['CLO PLO Results']
            
            # Apply formatting to main sheet
            _format_main_sheet(worksheet, df, sorted_clos, sorted_plos)
            
    except Exception as e:
        print(f"❌ Error occurred during Excel append: {str(e)}")
        raise e
    
    print(f"✅ Results appended to original file: {original_file_path}")
    return original_file_path


def _calculate_letter_grade(score):
    """Calculate letter grade based on numerical score using Habib University scale."""
    if score >= 95:
        return "A+"
    elif score >= 90:
        return "A"
    elif score >= 85:
        return "A-"
    elif score >= 80:
        return "B+"
    elif score >= 75:
        return "B"
    elif score >= 70:
        return "B-"
    elif score >= 67:
        return "C+"
    elif score >= 63:
        return "C"
    elif score >= 60:
        return "C-"
    else:
        return "F"


def _get_score_color(score):
    """Get color fill based on score value."""
    if score >= 70:
        return PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    elif score >= 60:
        return PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Yellow
    else:
        return PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Red


def _format_main_sheet(worksheet, df, sorted_clos, sorted_plos):
    """Apply formatting to the main results sheet."""
    
    # Format headers
    header_fill = PatternFill(start_color="6B2C91", end_color="6B2C91", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for column in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Format data cells with color coding
    for row in range(2, len(df) + 2):
        for column in range(2, len(df.columns)):  # Skip Overall Grade column for score coloring
            cell = worksheet.cell(row=row, column=column)
            value = cell.value
            
            if isinstance(value, (int, float)):
                cell.fill = _get_score_color(value)
                cell.alignment = Alignment(horizontal="center")
                
                # Special formatting for zero scores - keep red
                if value == 0:
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Red background
        
        # Format Overall Grade column differently
        grade_cell = worksheet.cell(row=row, column=len(df.columns))
        grade_cell.alignment = Alignment(horizontal="center")
        # You can add special formatting for grades if needed
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _create_summary_sheet(writer, clo_scores, plo_scores, grades, sorted_clos, sorted_plos):
    """
    Create and format the summary sheet with performance analytics.
    This function is kept for potential future use but is not currently called.
    """
    pass  # Function removed as summary sheet is no longer needed


def export_clo_plo_results(clo_scores, plo_scores, grades, data_dict, original_file_path):
    """
    Main export function - appends CLO/PLO results to the original uploaded file.
    
    Args:
        clo_scores (dict): CLO scores for all students
        plo_scores (dict): PLO scores for all students
        grades (dict): Final grades for all students (from terminal calculation)
        data_dict (dict): Original data from data.py
        original_file_path (str): Path to the original uploaded file
        
    Returns:
        str: Path to updated Excel file (same as original)
        
    Raises:
        Exception: If Excel modification fails
    """
    try:
        file_path = create_excel_output(clo_scores, plo_scores, grades, data_dict, original_file_path)
        print(f"✅ Results appended to original file: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"❌ Excel append failed: {str(e)}")
        raise e


if __name__ == "__main__":
    # Test function - can be used for standalone testing
    print("Excel Exporter Module - Modified to append to original files")
    print("This module should be imported, not run directly.")